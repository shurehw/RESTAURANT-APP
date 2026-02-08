"""
Excel Report Generator — multi-sheet openpyxl workbook for labor optimization.
"""

import json
import os
from datetime import datetime, date
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from ..db import get_db
from .. import config


# Style constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
DOW_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def _style_header_row(ws, row: int, max_col: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 30):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


class ExcelReportGenerator:
    """Generates multi-sheet Excel reports for labor optimization."""

    def __init__(self, venue_id: str, venue_name: str = None):
        self.venue_id = venue_id
        self.venue_name = venue_name or venue_id[:8]
        self.db = get_db()

    def generate_report(
        self,
        output_path: str = None,
        week_start: str = None,
    ) -> str:
        """
        Generate full 6-sheet Excel workbook.

        Args:
            output_path: Where to save the file
            week_start: Target week for staffing recs (YYYY-MM-DD Monday)

        Returns:
            Path to generated file
        """
        if not output_path:
            os.makedirs(config.REPORT_OUTPUT_DIR, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(
                config.REPORT_OUTPUT_DIR,
                f"labor_optimization_{self.venue_name}_{ts}.xlsx",
            )

        wb = Workbook()

        self._build_executive_summary(wb, week_start)
        self._build_dow_hour_profile(wb)
        self._build_staffing_recommendation(wb, week_start)
        self._build_backtest_results(wb)
        self._build_alerts(wb)
        self._build_cost_comparison(wb, week_start)

        # Remove default sheet if extra
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        wb.save(output_path)
        print(f"  [report] Saved to {output_path}")
        return output_path

    def _build_executive_summary(self, wb: Workbook, week_start: str = None):
        ws = wb.create_sheet("Executive Summary", 0)

        ws["A1"] = "Labor Optimization Report"
        ws["A1"].font = Font(bold=True, size=16, color="2F5496")
        ws["A2"] = f"Venue: {self.venue_name}"
        ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A4"] = f"Target Week: {week_start or 'N/A'}"

        row = 6
        ws.cell(row=row, column=1, value="Metric")
        ws.cell(row=row, column=2, value="Value")
        _style_header_row(ws, row, 2)

        # Load summary data
        profiles = self.db.select(
            "staffing_profiles", "day_of_week,hour_slot,p75_active_covers,servers_buffered",
            venue_id=f"eq.{self.venue_id}",
            order="profile_version.desc",
        )

        # Get latest version only
        if profiles:
            latest_v = max(p.get("profile_version", 1) for p in profiles if "profile_version" in p)
            profiles = [p for p in profiles if p.get("profile_version") == latest_v]

        # Key metrics
        metrics = []
        if profiles:
            all_p75 = [float(p["p75_active_covers"]) for p in profiles if p["p75_active_covers"]]
            peak_covers = max(all_p75) if all_p75 else 0
            avg_covers = sum(all_p75) / len(all_p75) if all_p75 else 0
            max_servers = max(int(p["servers_buffered"]) for p in profiles if p["servers_buffered"]) if profiles else 0

            metrics = [
                ("Peak Active Covers (P75)", f"{peak_covers:.0f}"),
                ("Avg Active Covers (P75)", f"{avg_covers:.0f}"),
                ("Max Servers (Buffered)", str(max_servers)),
                ("Profile Count", str(len(profiles))),
            ]

        # Backtest summary
        backtests = self.db.select(
            "backtest_results", "coverage_pct,wasted_labor_cost",
            venue_id=f"eq.{self.venue_id}",
            scenario="eq.buffered",
        )
        if backtests:
            avg_coverage = sum(float(b["coverage_pct"]) for b in backtests) / len(backtests)
            total_waste = sum(float(b["wasted_labor_cost"]) for b in backtests)
            metrics.extend([
                ("Backtest Coverage %", f"{avg_coverage:.1f}%"),
                ("Total Wasted Labor Cost", f"${total_waste:,.0f}"),
                ("Days Backtested", str(len(backtests))),
            ])

        # Unresolved alerts
        alerts = self.db.select(
            "staffing_alerts", "severity",
            venue_id=f"eq.{self.venue_id}",
            is_resolved="eq.false",
        )
        critical = sum(1 for a in alerts if a["severity"] == "critical")
        warning = sum(1 for a in alerts if a["severity"] == "warning")
        metrics.extend([
            ("Unresolved Alerts", str(len(alerts))),
            ("Critical Alerts", str(critical)),
        ])

        for i, (metric, value) in enumerate(metrics):
            r = row + 1 + i
            ws.cell(row=r, column=1, value=metric)
            ws.cell(row=r, column=2, value=value)
            ws.cell(row=r, column=1).border = THIN_BORDER
            ws.cell(row=r, column=2).border = THIN_BORDER

        _auto_width(ws)

    def _build_dow_hour_profile(self, wb: Workbook):
        ws = wb.create_sheet("DOW x Hour Profile")

        profiles = self.db.select(
            "staffing_profiles", "*",
            venue_id=f"eq.{self.venue_id}",
            order="profile_version.desc",
        )
        if not profiles:
            ws["A1"] = "No profiles available"
            return

        latest_v = max(p.get("profile_version", 1) for p in profiles)
        profiles = [p for p in profiles if p.get("profile_version") == latest_v]

        # Build grid: rows=hours, cols=DOW
        hours = sorted(set(p["hour_slot"] for p in profiles))
        dows = sorted(set(p["day_of_week"] for p in profiles))

        # Header
        ws.cell(row=1, column=1, value="P75 Active Covers")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)

        row = 3
        ws.cell(row=row, column=1, value="Hour")
        for i, dow in enumerate(dows):
            ws.cell(row=row, column=2 + i, value=DOW_NAMES[dow])
        _style_header_row(ws, row, 1 + len(dows))

        # Build lookup
        lookup = {}
        for p in profiles:
            lookup[(p["day_of_week"], p["hour_slot"])] = p

        max_covers = max(float(p.get("p75_active_covers", 0)) for p in profiles) or 1

        for h_idx, hour in enumerate(hours):
            r = row + 1 + h_idx
            ws.cell(row=r, column=1, value=f"{hour}:00")
            ws.cell(row=r, column=1).border = THIN_BORDER

            for d_idx, dow in enumerate(dows):
                profile = lookup.get((dow, hour))
                if profile:
                    val = float(profile.get("p75_active_covers", 0))
                    cell = ws.cell(row=r, column=2 + d_idx, value=round(val))
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal="center")
                    # Heatmap coloring
                    intensity = val / max_covers if max_covers else 0
                    if intensity > 0.8:
                        cell.fill = ALERT_FILL
                    elif intensity > 0.5:
                        cell.fill = WARN_FILL
                    elif intensity > 0.2:
                        cell.fill = GOOD_FILL

        _auto_width(ws)

    def _build_staffing_recommendation(self, wb: Workbook, week_start: str = None):
        ws = wb.create_sheet("Staffing Recommendation")

        if not week_start:
            ws["A1"] = "No target week specified"
            return

        forecasts = self.db.select(
            "daily_staffing_forecasts", "*",
            venue_id=f"eq.{self.venue_id}",
            forecast_date=f"gte.{week_start}",
            scenario="eq.buffered",
            order="forecast_date.asc",
            limit="7",
        )

        if not forecasts:
            ws["A1"] = f"No forecasts for week starting {week_start}"
            return

        row = 1
        ws.cell(row=row, column=1, value=f"Staffing Recommendation — Week of {week_start}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)

        row = 3
        headers = ["Date", "Day", "Hour", "Active Covers", "Servers", "Bartenders", "Seasonal"]
        for i, h in enumerate(headers):
            ws.cell(row=row, column=1 + i, value=h)
        _style_header_row(ws, row, len(headers))

        for forecast in forecasts:
            detail = json.loads(forecast["hourly_detail"]) if isinstance(forecast["hourly_detail"], str) else forecast["hourly_detail"]
            fd = forecast["forecast_date"]
            dow_name = DOW_NAMES[forecast["day_of_week"]]

            for entry in detail:
                row += 1
                ws.cell(row=row, column=1, value=fd)
                ws.cell(row=row, column=2, value=dow_name)
                ws.cell(row=row, column=3, value=f"{entry['hour']}:00")
                ws.cell(row=row, column=4, value=round(entry["active_covers"]))
                ws.cell(row=row, column=5, value=entry["servers"])
                ws.cell(row=row, column=6, value=entry["bartenders"])
                ws.cell(row=row, column=7, value=entry.get("seasonal_factor", 1.0))

                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).border = THIN_BORDER

        _auto_width(ws)

    def _build_backtest_results(self, wb: Workbook):
        ws = wb.create_sheet("Backtest Results")

        results = self.db.select(
            "backtest_results", "*",
            venue_id=f"eq.{self.venue_id}",
            scenario="eq.buffered",
            order="business_date.desc",
            limit="30",
        )

        if not results:
            ws["A1"] = "No backtest results available"
            return

        row = 1
        headers = ["Date", "Hours Analyzed", "Adequate", "Understaffed", "Overstaffed", "Coverage %", "Accuracy %", "Wasted Cost"]
        for i, h in enumerate(headers):
            ws.cell(row=row, column=1 + i, value=h)
        _style_header_row(ws, row, len(headers))

        for r_data in results:
            row += 1
            ws.cell(row=row, column=1, value=r_data["business_date"])
            ws.cell(row=row, column=2, value=r_data["hours_analyzed"])
            ws.cell(row=row, column=3, value=r_data["hours_adequate"])
            ws.cell(row=row, column=4, value=r_data["hours_understaffed"])
            ws.cell(row=row, column=5, value=r_data["hours_overstaffed"])
            ws.cell(row=row, column=6, value=float(r_data["coverage_pct"]))
            ws.cell(row=row, column=7, value=float(r_data.get("accuracy_pct", 0)))
            ws.cell(row=row, column=8, value=float(r_data["wasted_labor_cost"]))

            # Color coding
            coverage = float(r_data["coverage_pct"])
            cell_cov = ws.cell(row=row, column=6)
            if coverage >= 85:
                cell_cov.fill = GOOD_FILL
            elif coverage >= 70:
                cell_cov.fill = WARN_FILL
            else:
                cell_cov.fill = ALERT_FILL

            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER

        _auto_width(ws)

    def _build_alerts(self, wb: Workbook):
        ws = wb.create_sheet("Alerts")

        alerts = self.db.select(
            "staffing_alerts", "*",
            venue_id=f"eq.{self.venue_id}",
            is_resolved="eq.false",
            order="created_at.desc",
            limit="50",
        )

        if not alerts:
            ws["A1"] = "No active alerts"
            return

        row = 1
        headers = ["Date", "Hour", "Type", "Severity", "Message", "Actual Covers", "Rec. Servers"]
        for i, h in enumerate(headers):
            ws.cell(row=row, column=1 + i, value=h)
        _style_header_row(ws, row, len(headers))

        for alert in alerts:
            row += 1
            ws.cell(row=row, column=1, value=alert["alert_date"])
            ws.cell(row=row, column=2, value=f"{alert['hour_slot']}:00" if alert.get("hour_slot") is not None else "")
            ws.cell(row=row, column=3, value=alert["alert_type"])
            ws.cell(row=row, column=4, value=alert["severity"])
            ws.cell(row=row, column=5, value=alert["message"])
            ws.cell(row=row, column=6, value=alert.get("actual_covers", ""))
            ws.cell(row=row, column=7, value=alert.get("recommended_servers", ""))

            severity = alert["severity"]
            fill = ALERT_FILL if severity == "critical" else (WARN_FILL if severity == "warning" else None)
            if fill:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = fill

            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER

        _auto_width(ws)

    def _build_cost_comparison(self, wb: Workbook, week_start: str = None):
        ws = wb.create_sheet("Cost Comparison")

        if not week_start:
            ws["A1"] = "No target week specified for cost comparison"
            return

        # Get all three scenarios
        scenarios = {}
        for scenario in ["lean", "buffered", "safe"]:
            forecasts = self.db.select(
                "daily_staffing_forecasts",
                "forecast_date,total_labor_hours,estimated_labor_cost,estimated_covers,estimated_revenue",
                venue_id=f"eq.{self.venue_id}",
                forecast_date=f"gte.{week_start}",
                scenario=f"eq.{scenario}",
                order="forecast_date.asc",
                limit="7",
            )
            scenarios[scenario] = forecasts

        row = 1
        ws.cell(row=row, column=1, value=f"Cost Comparison — Week of {week_start}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)

        row = 3
        headers = ["Scenario", "Total Labor Hours", "Labor Cost", "Est. Covers", "Est. Revenue", "Labor %"]
        for i, h in enumerate(headers):
            ws.cell(row=row, column=1 + i, value=h)
        _style_header_row(ws, row, len(headers))

        for scenario_name in ["lean", "buffered", "safe"]:
            forecasts = scenarios.get(scenario_name, [])
            if not forecasts:
                continue

            total_hours = sum(float(f["total_labor_hours"]) for f in forecasts)
            total_cost = sum(float(f["estimated_labor_cost"]) for f in forecasts)
            total_covers = sum(int(f["estimated_covers"]) for f in forecasts)
            total_revenue = sum(float(f["estimated_revenue"]) for f in forecasts)
            labor_pct = (total_cost / total_revenue * 100) if total_revenue else 0

            row += 1
            ws.cell(row=row, column=1, value=scenario_name.title())
            ws.cell(row=row, column=2, value=round(total_hours, 1))
            ws.cell(row=row, column=3, value=round(total_cost, 2))
            ws.cell(row=row, column=3).number_format = '$#,##0'
            ws.cell(row=row, column=4, value=total_covers)
            ws.cell(row=row, column=5, value=round(total_revenue, 2))
            ws.cell(row=row, column=5).number_format = '$#,##0'
            ws.cell(row=row, column=6, value=round(labor_pct, 1))

            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER

        # Savings row
        if "lean" in scenarios and "safe" in scenarios and scenarios["lean"] and scenarios["safe"]:
            lean_cost = sum(float(f["estimated_labor_cost"]) for f in scenarios["lean"])
            safe_cost = sum(float(f["estimated_labor_cost"]) for f in scenarios["safe"])
            savings = safe_cost - lean_cost

            row += 2
            ws.cell(row=row, column=1, value="Potential Savings (Safe → Lean)")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=3, value=round(savings, 2))
            ws.cell(row=row, column=3).number_format = '$#,##0'
            ws.cell(row=row, column=3).font = Font(bold=True, color="006100")

        _auto_width(ws)
